"""
Open ChargePoint DataBase OCPDB
Copyright (C) 2021 binary butterfly GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

import logging
from datetime import datetime, timezone
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from celery.schedules import crontab
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from validataclass.exceptions import ValidationError
from validataclass.validators import DataclassValidator

from webapp.common.contexts import TelemetryContext
from webapp.common.error_handling.exceptions import RemoteException
from webapp.common.logging.models import LogMessageType
from webapp.models.source import Source, SourceStatus
from webapp.services.import_services.base_import_service import BaseImportService, SourceInfo

from .bnetza_excel_mapper import BnetzaExcelMapper
from .bnetza_excel_validators import BnetzaRowInput

logger = logging.getLogger(__name__)


class BnetzaExcelImportService(BaseImportService):
    schedule = crontab(day_of_month='1')

    bnetza_mapper: BnetzaExcelMapper = BnetzaExcelMapper()
    row_validator: DataclassValidator[BnetzaRowInput] = DataclassValidator(BnetzaRowInput)
    header_line: dict[str, str] = {
        'Betreiber': 'operator',
        'Anzeigename (Karte)': 'name_for_map',
        'Straße': 'address',
        'Hausnummer': 'housenumber',
        'Adresszusatz': 'additional_address_data',
        'Postleitzahl': 'postcode',
        'Ort': 'locality',
        'Kreis/kreisfreie Stadt': 'district',
        'Bundesland': 'land',
        'Breitengrad': 'lat',
        'Längengrad': 'lon',
        'Inbetriebnahmedatum': 'launch_date',
        'Nennleistung Ladeeinrichtung [kW]': 'connection_power',
        'Art der Ladeeinrichung': 'chargestation_type',
        'Anzahl Ladepunkte': 'connector_count',
        'Steckertypen1': 'connector_1_type',
        'P1 [kW]': 'connector_1_power',
        'Public Key1': 'connector_1_public_key',
        'Steckertypen2': 'connector_2_type',
        'P2 [kW]': 'connector_2_power',
        'Public Key2': 'connector_2_public_key',
        'Steckertypen3': 'connector_3_type',
        'P3 [kW]': 'connector_3_power',
        'Public Key3': 'connector_3_public_key',
        'Steckertypen4': 'connector_4_type',
        'P4 [kW]': 'connector_4_power',
        'Public Key4': 'connector_4_public_key',
        'Steckertypen5': 'connector_5_type',
        'P5 [kW]': 'connector_5_power',
        'Public Key5': 'connector_5_public_key',
        'Steckertypen6': 'connector_6_type',
        'P6 [kW]': 'connector_6_power',
        'Public Key6': 'connector_6_public_key',
    }
    source_info = SourceInfo(
        uid='bnetza_excel',
        name='Bundesnetzagentur',
        public_url='https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/Ladesaeulenkarte'
        '/start.html',
        source_url='https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/DL'
        '/Ladesaeuleninfrastruktur.xlsx',
        attribution_license='CC BY 4.0',
        attribution_url='https://creativecommons.org/licenses/by/4.0/deed.de',
        has_realtime_data=False,
    )

    def fetch_static_data(self):
        self.load_and_save_from_web()

    def load_and_save_from_web(self):
        source = self.get_source()
        static_data_updated_at = datetime.now(tz=timezone.utc)

        try:
            response = self.request()
        except RemoteException as e:
            logger.error(
                f'bnetza request failed: {e.to_dict()}',
                extra={'attributes': {'type': LogMessageType.IMPORT_SOURCE}},
            )
            self.update_source(source, static_status=SourceStatus.FAILED)
            return

        worksheet = load_workbook(filename=BytesIO(response.content)).active
        self.load_and_save(source=source, worksheet=worksheet, static_data_updated_at=static_data_updated_at)

    def load_and_save_from_file(self, import_file_path: Path):
        source = self.get_source()
        static_data_updated_at = datetime.now(tz=timezone.utc)

        try:
            worksheet = load_workbook(filename=import_file_path).active
        except InvalidFileException:
            logger.error(
                f'bnetza file {import_file_path} loading failed',
                extra={'attributes': {'type': LogMessageType.IMPORT_SOURCE}},
            )
            self.update_source(source, static_status=SourceStatus.FAILED)
            return

        self.load_and_save(source=source, worksheet=worksheet, static_data_updated_at=static_data_updated_at)

    def load_and_save(self, source: Source, worksheet: Worksheet, static_data_updated_at: datetime):
        try:
            self.check_mapping(worksheet[11])
        except ValidationError as e:
            logger.error(
                f'bnetza data has validation error: {e.to_dict()}',
                extra={'attributes': {'type': LogMessageType.IMPORT_SOURCE}},
            )
            self.update_source(
                source=source,
                static_status=SourceStatus.FAILED,
            )
            return

        rows_by_location_uid, static_error_count = self.get_rows_by_location_uid(worksheet=worksheet)
        static_success_count = 0

        location_updates = []
        for location_uid, rows in rows_by_location_uid.items():
            location_updates.append(self.bnetza_mapper.map_rows_to_location_update(location_uid, rows))
            static_success_count += 1

        self.save_location_updates(location_updates)

        self.update_source(
            source=source,
            static_error_count=static_error_count,
            static_status=SourceStatus.ACTIVE,
            static_data_updated_at=static_data_updated_at,
        )

        logger.info(
            f'Successfully updated {self.source_info.uid} static with {static_success_count} valid locations and '
            f'{static_error_count} failed locations.',
            extra={'attributes': {'type': LogMessageType.IMPORT_LOCATION}},
        )

    def check_mapping(self, row: tuple) -> None:
        if list(self.header_line.keys()) != [cell.value for cell in row]:
            raise ValidationError(reason='invalid xlsx header')

    def get_rows_by_location_uid(self, worksheet: Worksheet) -> tuple[dict[str, list[BnetzaRowInput]], int]:
        location_dict = {}
        static_error_count = 0
        # there are 10 rows of explanation over the header, plus 1 line header -> we start at row 12
        for table_row in worksheet.iter_rows(min_row=12):
            row_dict = {
                list(self.header_line.values())[i]: table_row[i].value for i in range(len(self.header_line.keys()))
            }

            # some postcodes are integer (why?! :( )
            row_dict['postcode'] = str(row_dict['postcode'])

            # normalize connectors
            for i in range(1, 7):
                connector_types = []
                if row_dict[f'connector_{i}_type']:
                    for item in row_dict[f'connector_{i}_type'].replace(';', ',').split(','):
                        connector_types.append(item.strip())
                row_dict[f'connector_{i}_type'] = connector_types

            try:
                row = self.row_validator.validate(row_dict)
                geo_hash = sha256(('%s-%s' % (row.lat, row.lon)).encode()).hexdigest()[:20]
                if geo_hash not in location_dict:
                    location_dict[geo_hash] = []
                location_dict[geo_hash].append(row)
            except ValidationError as e:
                logger.warning(
                    f'row {row_dict} is invalid: {e.to_dict()}',
                    extra={
                        'attributes': {
                            'type': LogMessageType.IMPORT_LOCATION,
                            TelemetryContext.LOCATION: row_dict.get('id'),
                        },
                    },
                )
                static_error_count += 1

        return location_dict, static_error_count
