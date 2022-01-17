# encoding: utf-8

"""
Open ChargePoint DataBase OCPDB
Copyright (C) 2021 binary butterfly GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

from hashlib import sha256
from typing import Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from validataclass.validators import DataclassValidator
from validataclass.exceptions import ValidationError
from webapp.services.external_helper import ExternalHelper
from webapp.services.base_service import BaseService
from webapp.services.generic.update_service import UpdateService
from webapp.repositories import LocationRepository, EvseRepository, ConnectorRepository
from .bnetza_validators import BnetzaRowInput
from .bnetza_mapper import BnetzaMapper


class MissingDataException(Exception):
    line: int
    reason: str

    def __init__(self, row: list, reason: str):
        self.row = row
        self.reason = reason


class BnetzaService(BaseService):
    update_service: UpdateService
    external_helper: ExternalHelper = ExternalHelper()
    bnetza_mapper: BnetzaMapper = BnetzaMapper()
    row_validator: DataclassValidator[BnetzaRowInput] = DataclassValidator(BnetzaRowInput)

    header_line = {
        'Betreiber': 'operator',
        'Straße': 'address',
        'Hausnummer': 'housenumber',
        'Adresszusatz': 'additional_address_data',
        'Postleitzahl': 'postcode',
        'Ort': 'locality',
        'Bundesland': 'land',
        'Kreis/kreisfreie Stadt': 'district',
        'Breitengrad': 'lat',
        'Längengrad': 'lon',
        'Inbetriebnahmedatum': 'launch_date',
        'Anschlussleistung': 'connection_power',
        'Art der Ladeeinrichung': 'chargestation_type',
        'Anzahl Ladepunkte': 'connector_count',
        'Steckertypen1': 'connector_1_type',
        'P1 [kW]': 'connector_1_power',
        'Public Key1': 'connector_1_public_key',
        'Steckertypen2': 'connector_2_type',
        'P2 [kW]': 'connector_2_power',
        'Public Key2': 'connector_2_public_key',
        'Steckertypen3': 'connector_3_type',
        'P3 [kW]': 'connector_3_power',
        'Public Key3': 'connector_3_public_key',
        'Steckertypen4': 'connector_4_type',
        'P4 [kW]': 'connector_4_power',
        'Public Key4': 'connector_4_public_key',
    }

    def __init__(
            self,
            *args,
            location_repository: LocationRepository,
            evse_repository: EvseRepository,
            connector_repository: ConnectorRepository,
            **kwargs):
        super().__init__(*args, **kwargs)
        self.update_service = UpdateService(
            logger=self.logger,
            config_helper=self.config_helper,
            location_repository=location_repository,
            evse_repository=evse_repository,
            connector_repository=connector_repository,
        )

    def load_and_save(self, import_file_path: str):
        worksheet = self.load_xlsx(import_file_path)
        self.check_mapping(worksheet[6])

        return self.save(
            self.load_data(
                worksheet=worksheet
            )
        )

    def load_xlsx(self, import_file_path: str) -> Worksheet:
        return load_workbook(filename=import_file_path).active

    def check_mapping(self, row: Tuple):
        if list(self.header_line.keys()) != [cell.value for cell in row]:
            raise Exception('invalid xlmx header')

    def load_data(self, worksheet: Worksheet) -> dict:
        location_dict = {}
        for table_row in worksheet.iter_rows(min_row=7):
            row_dict = {
                list(self.header_line.values())[i]: table_row[i].value
                for i in range(0, len(self.header_line.keys()))
            }
            for i in range(1, 5):
                row_dict['connector_%s_type' % i] = [
                    item.strip()
                    for item in row_dict['connector_%s_type' % i].replace(';', ',').split(',')
                    if item
                ]
            try:
                row = self.row_validator.validate(row_dict)
                geo_hash = sha256(('%s-%s' % (row.lat, row.lon)).encode()).hexdigest()[:20]
                if geo_hash not in location_dict:
                    location_dict[geo_hash] = []
                location_dict[geo_hash].append(row)
            except ValidationError as exception:
                print('%s: %s' % (row_dict, exception.to_dict()))
        return location_dict

    def save(self, location_dict: dict):
        for location_uid, rows in location_dict.items():
            self.update_service.upsert_location(
                self.bnetza_mapper.map_rows_to_location(location_uid, rows)
            )
